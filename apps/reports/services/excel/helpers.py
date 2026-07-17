from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(
    start_color="0B5EAC",
    end_color="0B5EAC",
    fill_type="solid"
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True
)

CENTER = Alignment(
    horizontal="center",
    vertical="center"
)

def style_header(ws, headers, row=1):

    for col_num, header in enumerate(headers, 1):

        cell = ws.cell(
            row=row,
            column=col_num,
            value=header
        )

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

def auto_size_columns(ws):

    for column_cells in ws.columns:

        length = 0

        column = column_cells[0].column

        for cell in column_cells:

            try:

                if len(str(cell.value)) > length:
                    length = len(str(cell.value))

            except:
                pass

        adjusted_width = length + 5

        ws.column_dimensions[
            get_column_letter(column)
        ].width = adjusted_width