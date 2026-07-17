from io import BytesIO
from django.http import HttpResponse
from django.db.models import Prefetch
from openpyxl import Workbook
from decimal import Decimal
from django.db.models import Sum, Count

from apps.sales.models import (
    SalePaymentMethod
)

from .helpers import (
    style_header,
    auto_size_columns
)


def generate_payment_methods_excel_report(
    *,
    company_id,
    branch_id=None,
    user_id=None,
    method=None,
    start_date=None,
    end_date=None,
    include_summary=True,
    detailed=True
):
    """
    Generate excel report for payment methods.

    Filters:
    - company
    - branch
    - user
    - payment method
    - date range
    """

    # ---------------------------------------------------
    # QUERYSET
    # ---------------------------------------------------

    queryset = SalePaymentMethod.objects.select_related(
        'company',
        'branch',
        'user',
        'sale',
        'payment'
    ).filter(
        company_id=company_id
    )

    # ---------------------------------------------------
    # FILTERS
    # ---------------------------------------------------

    if branch_id:
        queryset = queryset.filter(branch_id=branch_id)

    if user_id:
        queryset = queryset.filter(user_id=user_id)

    if method:
        queryset = queryset.filter(method=method)

    if start_date:
        queryset = queryset.filter(date__gte=start_date)

    if end_date:
        queryset = queryset.filter(date__lte=end_date)

    queryset = queryset.order_by('-date')

    # ---------------------------------------------------
    # CREATE EXCEL
    # ---------------------------------------------------

    wb = Workbook()
    ws = wb.active
    ws.title = "Payment Methods"

    headers = [
        "Date",
        "Company",
        "Branch",
        "User",
        "Method",
        "Amount",
        "Comment"
    ]

    if detailed:
        headers.extend([
            "Sale Reference",
            "Sale Status",
            "Payment ID"
        ])

    # HEADER STYLE
    style_header(ws, headers)

    # ---------------------------------------------------
    # DATA
    # ---------------------------------------------------

    row = 2

    total_amount = 0

    for payment_method in queryset:

        total_amount += float(payment_method.amount)

        data = [
            payment_method.date.strftime("%Y-%m-%d %H:%M")
            if payment_method.date else "",

            payment_method.company.company_name
            if payment_method.company else "",

            payment_method.branch.name_branch
            if payment_method.branch else "",

            payment_method.user.username
            if payment_method.user else "",

            payment_method.method,

            float(payment_method.amount),

            payment_method.comment or ""
        ]

        if detailed:
            data.extend([
                payment_method.sale.reference
                if payment_method.sale else "",

                payment_method.sale.status
                if payment_method.sale else "",

                payment_method.payment.id
                if payment_method.payment else ""
            ])

        for col_num, value in enumerate(data, start=1):
            ws.cell(row=row, column=col_num, value=value)

        row += 1

    # ---------------------------------------------------
    # SUMMARY
    # ---------------------------------------------------

    if include_summary:

        row += 2

        ws.cell(row=row, column=1, value="SUMMARY")
        row += 1

        summary_headers = [
            "Method",
            "Transactions",
            "Total"
        ]

        for col_num, header in enumerate(summary_headers, start=1):
            ws.cell(row=row, column=col_num, value=header)

        row += 1

        summary_queryset = queryset.values(
            'method'
        ).annotate(
            total=Sum('amount'),
            transactions=Count('id')
        )

        for item in summary_queryset:

            summary_data = [
                item['method'],
                item['transactions'],
                float(item['total'] or 0)
            ]

            for col_num, value in enumerate(summary_data, start=1):
                ws.cell(row=row, column=col_num, value=value)

            row += 1

        # GLOBAL TOTAL
        row += 1

        ws.cell(row=row, column=1, value="GLOBAL TOTAL")
        ws.cell(row=row, column=2, value=float(total_amount))

    # ---------------------------------------------------
    # EXCEL SETTINGS
    # ---------------------------------------------------

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    auto_size_columns(ws)

    # ---------------------------------------------------
    # SAVE MEMORY
    # ---------------------------------------------------

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # ---------------------------------------------------
    # RESPONSE
    # ---------------------------------------------------

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response[
        'Content-Disposition'
    ] = 'attachment; filename="payment_methods_report.xlsx"'

    return response






from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image,
    PageBreak
)
from collections import defaultdict
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch

#THIS IS VERY IMPORT BECAUSE THIS IS FOR THAT <matplotlib> NOT USE INTERFACE
#IF NOT HAVE THIS AFTER OF <matplotlib.pyplot> THE SERVER WILL ARE BROKEN
import matplotlib
matplotlib.use('Agg')

import matplotlib.pyplot as plt
def generate_payment_methods_pdf_report(
    *,
    company_id,
    branch_id=None,
    user_id=None,
    method=None,
    start_date=None,
    end_date=None,
    include_summary=True,
    detailed=True
):
    """
    Generate PDF report for payment methods
    with charts and summary.
    """

    # =====================================================
    # QUERYSET
    # =====================================================

    queryset = SalePaymentMethod.objects.select_related(
        'company',
        'branch',
        'user',
        'sale',
        'payment'
    ).filter(
        company_id=company_id
    )

    # =====================================================
    # FILTERS
    # =====================================================

    if branch_id:
        queryset = queryset.filter(branch_id=branch_id)

    if user_id:
        queryset = queryset.filter(user_id=user_id)

    if method:
        queryset = queryset.filter(method=method)

    if start_date:
        queryset = queryset.filter(date__gte=start_date)

    if end_date:
        queryset = queryset.filter(date__lte=end_date)

    queryset = queryset.order_by('-date')

    # =====================================================
    # SUMMARY DATA
    # =====================================================

    summary_queryset = queryset.values(
        'method'
    ).annotate(
        total=Sum('amount'),
        transactions=Count('id')
    )

    total_amount = Decimal('0')

    methods = []
    totals = []

    for item in summary_queryset:

        method_name = item['method'] or "Unknown"
        amount = float(item['total'] or 0)

        methods.append(method_name)
        totals.append(amount)

        total_amount += Decimal(amount)

    # =====================================================
    # GENERATE CHART
    # =====================================================

    chart_buffer = BytesIO()

    fig, ax = plt.subplots(figsize=(8, 4))

    ax.bar(methods, totals)

    ax.set_title("Payment Methods Distribution")
    ax.set_ylabel("Amount")

    plt.xticks(rotation=20)

    plt.tight_layout()

    fig.savefig(chart_buffer, format='png')

    plt.close(fig)

    chart_buffer.seek(0)

    # =====================================================
    # PDF CONFIG
    # =====================================================

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=30,
        leftMargin=30,
        topMargin=50,
        bottomMargin=50
    )

    styles = getSampleStyleSheet()

    elements = []

    # =====================================================
    # TITLE
    # =====================================================

    title = Paragraph(
        "<font size=20><b>Payment Methods Report</b></font>",
        styles['Title']
    )

    elements.append(title)

    elements.append(Spacer(1, 20))

    # =====================================================
    # CHART
    # =====================================================

    chart = Image(
        chart_buffer,
        width=6.5 * inch,
        height=3.2 * inch
    )

    elements.append(chart)

    elements.append(Spacer(1, 20))

    # =====================================================
    # TABLE DATA
    # =====================================================

    data = [[
        "Date",
        "Branch",
        "User",
        "Method",
        "Amount",
        "Comment"
    ]]

    if detailed:
        data[0].extend([
            "Sale Ref",
            "Sale Status"
        ])

    # =====================================================
    # TABLE ROWS
    # =====================================================

    for payment_method in queryset.iterator(chunk_size=500):

        row = [

            payment_method.date.strftime("%Y-%m-%d %H:%M")
            if payment_method.date else "",

            payment_method.branch.name_branch
            if payment_method.branch else "",

            payment_method.user.username
            if payment_method.user else "",

            payment_method.method,

            f"${float(payment_method.amount):,.2f}",

            payment_method.comment or ""
        ]

        if detailed:

            row.extend([

                payment_method.sale.reference
                if payment_method.sale else "",

                payment_method.sale.status
                if payment_method.sale else ""
            ])

        data.append(row)

    # =====================================================
    # TABLE
    # =====================================================

    table = Table(
        data,
        repeatRows=1
    )

    table.setStyle(TableStyle([

        # HEADER
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0B5EAC")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),

        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),

        ('FONTSIZE', (0, 0), (-1, -1), 8),

        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),

        # BODY
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),

        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),

        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

        ('ALIGN', (4, 1), (4, -1), 'RIGHT'),

    ]))

    elements.append(table)

    elements.append(Spacer(1, 25))

    # =====================================================
    # SUMMARY SECTION
    # =====================================================

    if include_summary:

        summary_title = Paragraph(
            "<b>Summary</b>",
            styles['Heading2']
        )

        elements.append(summary_title)

        elements.append(Spacer(1, 10))

        summary_data = [[
            "Method",
            "Transactions",
            "Total"
        ]]

        for item in summary_queryset:

            summary_data.append([

                item['method'],

                item['transactions'],

                f"${float(item['total'] or 0):,.2f}"
            ])

        summary_data.append([
            "",
            "GLOBAL TOTAL",
            f"${float(total_amount):,.2f}"
        ])

        summary_table = Table(summary_data)

        summary_table.setStyle(TableStyle([

            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0B5EAC")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),

            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),

            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),

            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#EAF2FF")),

            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),

        ]))

        elements.append(summary_table)

    # =====================================================
    # FOOTER
    # =====================================================

    def add_footer(canvas, doc):

        canvas.saveState()

        canvas.setFont('Helvetica', 9)

        canvas.setFillColor(colors.grey)

        canvas.drawString(
            30,
            20,
            "Powered by www.Denty.Cloud"
        )

        canvas.drawRightString(
            580,
            20,
            f"Page {doc.page}"
        )

        canvas.restoreState()

    # =====================================================
    # BUILD PDF
    # =====================================================

    doc.build(
        elements,
        onFirstPage=add_footer,
        onLaterPages=add_footer
    )

    pdf = buffer.getvalue()

    buffer.close()

    # =====================================================
    # RESPONSE
    # =====================================================

    response = HttpResponse(
        content_type='application/pdf'
    )

    response[
        'Content-Disposition'
    ] = 'attachment; filename="payment_methods_report.pdf"'

    response.write(pdf)

    return response

