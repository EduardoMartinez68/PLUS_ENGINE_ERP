from io import BytesIO
from django.http import HttpResponse
from django.db.models import Prefetch
from openpyxl import Workbook
from decimal import Decimal

from apps.sales.models import (
    Sale,
    SaleItem,
    SaleHistory,
    SalePaymentMethod
)

from .helpers import (
    style_header,
    auto_size_columns
)


def generate_sales_excel_report(
    *,
    company_id,
    branch_id=None,
    user_id=None,
    customer_id=None,
    start_date=None,
    end_date=None,
    status=None,
    include_items=False,
    include_payments=False,
    include_payment_methods=False,
    detailed=False
):
    """
    Generate dynamic excel report for sales
    """

    queryset = Sale.objects.select_related(
        'customer',
        'branch',
        'user',
        'company'
    )

    # FILTERS
    queryset = queryset.filter(company_id=company_id)

    if branch_id:
        queryset = queryset.filter(branch_id=branch_id)

    if user_id:
        queryset = queryset.filter(user_id=user_id)

    if customer_id:
        queryset = queryset.filter(customer_id=customer_id)

    if status:
        queryset = queryset.filter(status=status)

    if start_date:
        queryset = queryset.filter(creationDate__gte=start_date)

    if end_date:
        queryset = queryset.filter(creationDate__lte=end_date)

    # PREFETCH OPTIMIZATION
    if include_items:
        queryset = queryset.prefetch_related(
            Prefetch('items', queryset=SaleItem.objects.all())
        )

    if include_payments:
        queryset = queryset.prefetch_related(
            Prefetch('payments', queryset=SaleHistory.objects.all())
        )

    if include_payment_methods:
        queryset = queryset.prefetch_related(
            Prefetch(
                'payment_methods',
                queryset=SalePaymentMethod.objects.all()
            )
        )

    queryset = queryset.order_by('-creationDate')

    # CREATE EXCEL
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    headers = [
        "Reference",
        "Customer",
        "Branch",
        "User",
        "Status",
        "Subtotal",
        "Discount",
        "Taxes",
        "Total",
        "Paid",
        "Balance",
        "Currency",
        "Creation Date"
    ]

    if detailed:
        headers.extend([
            "Expiration Date",
            "Fiscal Status",
            "UUID"
        ])

    style_header(ws, headers)

    # DATA
    row = 2

    for sale in queryset:

        data = [
            sale.reference,
            sale.customer.name if sale.customer else "",
            sale.branch.name_branch if sale.branch else "",
            sale.user.username if sale.user else "",
            sale.status,
            Decimal(sale.subtotal),
            Decimal(sale.discount_total),
            Decimal(sale.tax_total),
            Decimal(sale.total),
            Decimal(sale.amount_paid),
            Decimal(sale.balance),
            sale.currency,
            sale.creationDate.strftime("%Y-%m-%d %H:%M")
        ]

        if detailed:
            data.extend([
                sale.expiration_date.strftime("%Y-%m-%d")
                if sale.expiration_date else "",
                sale.fiscal_status,
                sale.fiscal_uuid
            ])

        for col_num, value in enumerate(data, 1):
            ws.cell(row=row, column=col_num, value=value)

        row += 1

        # ITEMS
        if include_items:

            ws.cell(row=row, column=1, value="ITEMS")
            row += 1

            item_headers = [
                "Name",
                "Qty",
                "Unit Price",
                "Discount",
                "Tax",
                "Subtotal",
                "Total"
            ]

            for col_num, header in enumerate(item_headers, 1):
                ws.cell(row=row, column=col_num, value=header)

            row += 1

            for item in sale.items.all():

                item_data = [
                    item.name,
                    Decimal(item.quantity),
                    Decimal(item.unit_price),
                    Decimal(item.discount),
                    Decimal(item.tax_amount),
                    Decimal(item.subtotal),
                    Decimal(item.total)
                ]

                for col_num, value in enumerate(item_data, 1):
                    ws.cell(row=row, column=col_num, value=value)

                row += 1

        # PAYMENTS
        if include_payments:

            row += 1
            ws.cell(row=row, column=1, value="PAYMENTS")
            row += 1

            payment_headers = [
                "Cash Received",
                "Change",
                "Old Balance",
                "New Balance",
                "Date"
            ]

            for col_num, header in enumerate(payment_headers, 1):
                ws.cell(row=row, column=col_num, value=header)

            row += 1

            for payment in sale.payments.all():

                payment_data = [
                    Decimal(payment.cash_received),
                    Decimal(payment.change_given),
                    Decimal(payment.old_balance),
                    Decimal(payment.new_balance),
                    payment.date.strftime("%Y-%m-%d %H:%M")
                ]

                for col_num, value in enumerate(payment_data, 1):
                    ws.cell(row=row, column=col_num, value=value)

                row += 1

        row += 2

    auto_size_columns(ws)

    # SAVE MEMORY
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response[
        'Content-Disposition'
    ] = 'attachment; filename="sales_report.xlsx"'

    return response




from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image,
    PageBreak
)
from collections import defaultdict
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch

#THIS IS VERY IMPORT BECAUSE THIS IS FOR THAT <matplotlib> NOT USE INTERFACE
#IF NOT HAVE THIS AFTER OF <matplotlib.pyplot> THE SERVER WILL ARE BROKEN
import matplotlib
matplotlib.use('Agg')

import matplotlib.pyplot as plt
def generate_sales_pdf_report(
    *,
    company_id,
    branch_id=None,
    user_id=None,
    customer_id=None,
    start_date=None,
    end_date=None,
    status=None,
    include_items=False
):

    queryset = Sale.objects.select_related(
        'customer',
        'branch',
        'user',
        'company'
    )

    queryset = queryset.filter(company_id=company_id)

    if branch_id:
        queryset = queryset.filter(branch_id=branch_id)

    if user_id:
        queryset = queryset.filter(user_id=user_id)

    if customer_id:
        queryset = queryset.filter(customer_id=customer_id)

    if status:
        queryset = queryset.filter(status=status)

    if start_date:
        queryset = queryset.filter(creationDate__gte=start_date)

    if end_date:
        queryset = queryset.filter(creationDate__lte=end_date)

    queryset = queryset.order_by('-creationDate')

    # =========================================================
    # GENERATE CHART
    # =========================================================

    sales_by_day = defaultdict(float)

    for sale in queryset:
        day = sale.creationDate.strftime("%Y-%m-%d")
        sales_by_day[day] += float(sale.total)

    chart_buffer = BytesIO()

    plt.figure(figsize=(8, 4))

    plt.plot(
        list(sales_by_day.keys()),
        list(sales_by_day.values())
    )

    plt.xticks(rotation=45)
    plt.title("Sales by Day")
    plt.tight_layout()

    plt.savefig(chart_buffer, format='png')
    plt.close()

    chart_buffer.seek(0)

    # =========================================================
    # PDF
    # =========================================================

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=60,
        bottomMargin=60
    )

    styles = getSampleStyleSheet()

    elements = []

    # =========================================================
    # TITLE
    # =========================================================

    title = Paragraph(
        "<b>Sales Report</b>",
        styles['Title']
    )

    elements.append(title)
    elements.append(Spacer(1, 20))

    # =========================================================
    # CHART
    # =========================================================

    chart = Image(chart_buffer, width=6*inch, height=3*inch)

    elements.append(chart)
    elements.append(Spacer(1, 20))

    # =========================================================
    # TABLE
    # =========================================================

    data = [[
        "Reference",
        "Customer",
        "Branch",
        "User",
        "Status",
        "Total",
        "Paid",
        "Balance",
        "Date"
    ]]

    total_sales = Decimal('0')

    for sale in queryset:

        total_sales += Decimal(sale.total)

        data.append([
            sale.reference,
            sale.customer.name if sale.customer else "",
            sale.branch.name_branch if sale.branch else "",
            sale.user.username if sale.user else "",
            sale.status,
            f"${sale.total}",
            f"${sale.amount_paid}",
            f"${sale.balance}",
            sale.creationDate.strftime("%Y-%m-%d")
        ])

    table = Table(data, repeatRows=1)

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0B5EAC")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),

        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),

        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

        ('BACKGROUND', (0, 1), (-1, -1), colors.white),

        ('GRID', (0, 0), (-1, -1), 1, colors.grey),

        ('ALIGN', (5, 1), (-1, -1), 'CENTER'),
    ]))

    elements.append(table)

    elements.append(Spacer(1, 20))

    # =========================================================
    # TOTALS
    # =========================================================

    total_paragraph = Paragraph(
        f"<b>Total Sales:</b> ${total_sales}",
        styles['Heading2']
    )

    elements.append(total_paragraph)

    # =========================================================
    # FOOTER
    # =========================================================

    def add_footer(canvas, doc):

        canvas.saveState()

        footer_text = "Powered by www.Denty.Cloud"

        canvas.setFont('Helvetica', 9)

        canvas.drawString(
            40,
            20,
            footer_text
        )

        canvas.drawRightString(
            570,
            20,
            f"Page {doc.page}"
        )

        canvas.restoreState()

    # =========================================================
    # BUILD PDF
    # =========================================================

    doc.build(
        elements,
        onFirstPage=add_footer,
        onLaterPages=add_footer
    )

    pdf = buffer.getvalue()

    buffer.close()

    response = HttpResponse(
        content_type='application/pdf'
    )

    response['Content-Disposition'] = (
        'attachment; filename="sales_report.pdf"'
    )

    response.write(pdf)

    return response