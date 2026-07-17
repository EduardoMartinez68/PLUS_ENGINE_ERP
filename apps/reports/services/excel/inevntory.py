from io import BytesIO
from decimal import Decimal

from django.http import HttpResponse
from django.db.models import Sum
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

from apps.services.models import (
    Inventory,
    HistoryInventory
)

from .helpers import (
    style_header,
    auto_size_columns
)

# =========================================================
# EXPORT CURRENT INVENTORY
# =========================================================

def generate_inventory_excel_report(
    *,
    company_id,
    branch_id=None,
    only_low_stock=False
):
    """
    Export current inventory report
    """

    queryset = Inventory.objects.select_related(
        'company',
        'branch',
        'pack',
        'pack__category',
        'pack__department'
    ).filter(
        company_id=company_id
    )

    # -------------------------------------------------
    # FILTERS
    # -------------------------------------------------

    if branch_id:
        queryset = queryset.filter(branch_id=branch_id)

    if only_low_stock:
        queryset = queryset.filter(
            stock__lte=models.F('min_stock')
        )

    queryset = queryset.order_by('pack__name')

    # -------------------------------------------------
    # CREATE EXCEL
    # -------------------------------------------------

    wb = Workbook()
    ws = wb.active
    ws.title = "Current Inventory"

    headers = [
        "Branch",
        "Product",
        "Category",
        "Department",
        "Stock",
        "Min Stock",
        "Difference",
        "Status",
        "Track Inventory",
        "Activated",
        "Last Update"
    ]

    style_header(ws, headers)

    # -------------------------------------------------
    # COLORS
    # -------------------------------------------------

    red_fill = PatternFill(
        start_color="FFCCCC",
        end_color="FFCCCC",
        fill_type="solid"
    )

    green_fill = PatternFill(
        start_color="CCFFCC",
        end_color="CCFFCC",
        fill_type="solid"
    )

    warning_fill = PatternFill(
        start_color="FFF4CC",
        end_color="FFF4CC",
        fill_type="solid"
    )

    # -------------------------------------------------
    # DATA
    # -------------------------------------------------

    row = 2

    total_stock = Decimal('0')

    for inventory in queryset:

        difference = inventory.stock - inventory.min_stock

        total_stock += inventory.stock

        # Determine inventory status
        if inventory.stock <= 0:
            status = "OUT OF STOCK"
        elif inventory.stock <= inventory.min_stock:
            status = "LOW STOCK"
        else:
            status = "OK"

        data = [
            inventory.branch.name_branch if inventory.branch else "",
            inventory.pack.name if inventory.pack else "",
            inventory.pack.category.name
            if inventory.pack and inventory.pack.category else "",

            inventory.pack.department.name
            if inventory.pack and inventory.pack.department else "",

            float(inventory.stock),
            float(inventory.min_stock),
            float(difference),
            status,

            "YES" if inventory.pack.track_inventory else "NO",
            "YES" if inventory.pack.activated else "NO",

            inventory.last_update.strftime("%Y-%m-%d %H:%M")
            if inventory.last_update else ""
        ]

        for col_num, value in enumerate(data, start=1):

            cell = ws.cell(
                row=row,
                column=col_num,
                value=value
            )

            # APPLY COLORS TO STATUS
            if col_num == 8:

                if status == "OUT OF STOCK":
                    cell.fill = red_fill

                elif status == "LOW STOCK":
                    cell.fill = warning_fill

                else:
                    cell.fill = green_fill

        row += 1

    # -------------------------------------------------
    # SUMMARY
    # -------------------------------------------------

    row += 2

    ws.cell(row=row, column=1, value="TOTAL PRODUCTS")
    ws.cell(row=row, column=2, value=queryset.count())

    row += 1

    ws.cell(row=row, column=1, value="TOTAL STOCK")
    ws.cell(row=row, column=2, value=float(total_stock))

    # -------------------------------------------------
    # SETTINGS
    # -------------------------------------------------

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    auto_size_columns(ws)

    # -------------------------------------------------
    # SAVE MEMORY
    # -------------------------------------------------

    buffer = BytesIO()

    wb.save(buffer)

    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response[
        'Content-Disposition'
    ] = 'attachment; filename="inventory_report.xlsx"'

    return response


# =========================================================
# EXPORT INVENTORY HISTORY
# =========================================================

def generate_inventory_history_excel_report(
    *,
    company_id,
    branch_id=None,
    pack_id=None,
    unity_type=None,
    user_id=None,
    start_date=None,
    end_date=None
):
    """
    Export inventory movement history
    """

    queryset = HistoryInventory.objects.select_related(
        'company',
        'branch',
        'pack',
        'user'
    ).filter(
        company_id=company_id
    )

    # -------------------------------------------------
    # FILTERS
    # -------------------------------------------------

    if branch_id:
        queryset = queryset.filter(branch_id=branch_id)

    if pack_id:
        queryset = queryset.filter(pack_id=pack_id)

    if unity_type and unity_type!='all':
        queryset = queryset.filter(unity_type=unity_type)

    if user_id:
        queryset = queryset.filter(user_id=user_id)

    if start_date:
        queryset = queryset.filter(update_date__gte=start_date)

    if end_date:
        queryset = queryset.filter(update_date__lte=end_date)

    queryset = queryset.order_by('-update_date')

    # -------------------------------------------------
    # CREATE EXCEL
    # -------------------------------------------------

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory History"

    headers = [
        "Date",
        "Branch",
        "Product",
        "User",
        "Unity Type",
        "Old Stock",
        "New Stock",
        "Difference",
        "Movement Type"
    ]

    style_header(ws, headers)

    # -------------------------------------------------
    # COLORS
    # -------------------------------------------------

    green_fill = PatternFill(
        start_color="CCFFCC",
        end_color="CCFFCC",
        fill_type="solid"
    )

    red_fill = PatternFill(
        start_color="FFCCCC",
        end_color="FFCCCC",
        fill_type="solid"
    )

    gray_fill = PatternFill(
        start_color="E0E0E0",
        end_color="E0E0E0",
        fill_type="solid"
    )

    # -------------------------------------------------
    # DATA
    # -------------------------------------------------

    row = 2

    total_entries = Decimal('0')
    total_outputs = Decimal('0')

    for history in queryset:

        difference = history.new_stock - history.old_stock

        # Determine movement type
        if difference > 0:
            movement_type = "ENTRY"
            total_entries += difference

        elif difference < 0:
            movement_type = "OUTPUT"
            total_outputs += abs(difference)

        else:
            movement_type = "NO CHANGE"

        data = [
            history.update_date.strftime("%Y-%m-%d %H:%M")
            if history.update_date else "",

            history.branch.name_branch if history.branch else "",

            history.pack.name if history.pack else "",

            history.user.username if history.user else "",

            history.unity_type,

            float(history.old_stock),
            float(history.new_stock),
            float(difference),
            movement_type
        ]

        for col_num, value in enumerate(data, start=1):

            cell = ws.cell(
                row=row,
                column=col_num,
                value=value
            )

            # COLOR MOVEMENTS
            if col_num == 9:

                if movement_type == "ENTRY":
                    cell.fill = green_fill

                elif movement_type == "OUTPUT":
                    cell.fill = red_fill

                else:
                    cell.fill = gray_fill

        row += 1

    # -------------------------------------------------
    # SUMMARY
    # -------------------------------------------------

    row += 2

    ws.cell(row=row, column=1, value="TOTAL ENTRIES")
    ws.cell(row=row, column=2, value=float(total_entries))

    row += 1

    ws.cell(row=row, column=1, value="TOTAL OUTPUTS")
    ws.cell(row=row, column=2, value=float(total_outputs))

    row += 1

    ws.cell(row=row, column=1, value="TOTAL MOVEMENTS")
    ws.cell(row=row, column=2, value=queryset.count())

    # -------------------------------------------------
    # SETTINGS
    # -------------------------------------------------

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    auto_size_columns(ws)

    # -------------------------------------------------
    # SAVE MEMORY
    # -------------------------------------------------

    buffer = BytesIO()

    wb.save(buffer)

    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response[
        'Content-Disposition'
    ] = 'attachment; filename="inventory_history_report.xlsx"'

    return response