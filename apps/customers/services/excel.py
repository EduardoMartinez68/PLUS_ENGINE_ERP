import openpyxl
from openpyxl import Workbook
import tempfile
from django.http import HttpResponse
from openpyxl import Workbook
from io import BytesIO
import os 
import json
from django.core.exceptions import ValidationError
from apps.customers.models import Customer, CustomerType, CustomerSource
from openpyxl.styles import PatternFill, Font
from ..plus_wrapper import Plus


def create_excel(user):
    '''
    this script is for create a excel and send to the frontend for that the 
    user can upload most speed the information of his customers
    '''
    language=user.language

    #here we will to create the book of excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Plus"

    #Now we will to create the header of excel in the language of the user
    headers = get_header_of_excel(language)

    #Styles for the header
    header_fill = PatternFill(start_color="4183C1", end_color="4F81BD", fill_type="solid")  
    header_font = Font(color="FFFFFF", bold=True)  # letter white and black

    # add the header and apply the styles
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font

    #save in memory the excel
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Prepare the HTTP response for download
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="customers.xlsx"'
    return response

def get_header_of_excel(language)->list:
    '''
    return the header of the excel for upload customer
    with help of the language that need the user
    '''
    #get the path of the file of translate 
    translate_data=get_file_translate(language)

    #this be the key of the name that exist in the <translate.json>
    EXCEL_KEYS = [
        "customers.excel.sku",
        "customers.excel.name",
        "customers.excel.email",
        "customers.excel.phone",
        "customers.excel.cellphone",
        "customers.excel.date_of_birth",
        "customers.excel.gender",
        "customers.excel.country",
        "customers.excel.address",
        "customers.excel.city",
        "customers.excel.state",
        "customers.excel.postal_code",
        "customers.excel.num_ext",
        "customers.excel.num_int",
        "customers.excel.reference",
        "customers.excel.is_company",
        "customers.excel.company_name",
        "customers.excel.contact_name",
        "customers.excel.contact_email",
        "customers.excel.contact_phone",
        "customers.excel.contact_cellphone",
        "customers.excel.website",
        "customers.excel.note",
        "customers.excel.points",
        "customers.excel.credit",
        "customers.excel.tags",
        "customers.excel.customer_type",
        "customers.excel.customer_source",
        "customers.excel.priority",
        "customers.excel.price_list_number"
    ]



    #return the language that the user need, if not exist we will to return the Spanish for default
    headers_list = [translate_data.get(key) for key in EXCEL_KEYS]
    return headers_list

def get_file_translate(language)->list:
    #create the path of the file <translate.json> for read the translate
    current_dir = os.path.dirname(__file__) #get the path current of the script
    file_path = os.path.join(current_dir, '..', 'locale', language, 'translate.json') #We get the path where the translation file is located.

    #now we will see if this file exist and if not exist we will to return all the language in spanish
    if not os.path.exists(file_path):
        return  {
        "customers.excel.sku": "Clave",
        "customers.excel.name": "Nombre *",
        "customers.excel.email": "Correo",
        "customers.excel.phone": "Teléfono",
        "customers.excel.cellphone": "Celular",
        "customers.excel.date_of_birth": "Fecha de Nacimiento",
        "customers.excel.gender": "Género",
        "customers.excel.country": "País",
        "customers.excel.address": "Dirección",
        "customers.excel.city": "Ciudad",
        "customers.excel.state": "Estado",
        "customers.excel.postal_code": "Código Postal",
        "customers.excel.num_ext": "Número Exterior",
        "customers.excel.num_int": "Número Interior",
        "customers.excel.reference": "Referencia",
        "customers.excel.is_company": "¿Es Empresa?",
        "customers.excel.company_name": "Nombre de la Empresa",
        "customers.excel.contact_name": "Nombre del Contacto",
        "customers.excel.contact_email": "Correo del Contacto",
        "customers.excel.contact_phone": "Teléfono del Contacto",
        "customers.excel.contact_cellphone": "Celular del Contacto",
        "customers.excel.website": "Sitio Web",
        "customers.excel.note": "Nota",
        "customers.excel.points": "Puntos",
        "customers.excel.credit": "Crédito",
        "customers.excel.tags": "Etiquetas (JSON o separadas por coma)",
        "customers.excel.customer_type": "Tipo de Cliente (ID o Nombre)",
        "customers.excel.customer_source": "Fuente del Cliente (ID o Nombre)",
        "customers.excel.priority": "Prioridad",
        "customers.excel.price_list_number": "Número de Lista de Precio"
    }


    #if exist now we will read all the information of the json for after search the information of translate
    with open(file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    return data

def get_the_data_of_the_file_excel(excel_file)->list:
    try:
        #here we will to read the file
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        data = []

        #get the information of the file with help of his key
        for row in sheet.iter_rows(min_row=2, values_only=True):

            #save the information in a list of data fro after save in the database
            data.append({
                "name": row[0] if len(row) > 0 else None,
                "email": row[1] if len(row) > 1 else None,
                "phone": row[2] if len(row) > 2 else None,
                "cellphone": row[3] if len(row) > 3 else None,
                "date_of_birth": row[4] if len(row) > 4 else None,
                "gender": row[5] if len(row) > 5 else None,
                "country": row[6] if len(row) > 6 else None,
                "address": row[7] if len(row) > 7 else None,
                "city": row[8] if len(row) > 8 else None,
                "state": row[9] if len(row) > 9 else None,
                "postal_code": row[10] if len(row) > 10 else None,
                "num_ext": row[11] if len(row) > 11 else None,
                "num_int": row[12] if len(row) > 12 else None,
                "reference": row[13] if len(row) > 13 else None,
                "this_customer_is_a_company": row[14] if len(row) > 14 else None,
                "company_name": row[15] if len(row) > 15 else None,
                "contact_name": row[16] if len(row) > 16 else None,
                "contact_email": row[17] if len(row) > 17 else None,
                "contact_phone": row[18] if len(row) > 18 else None,
                "contact_cellphone": row[19] if len(row) > 19 else None,
                "website": row[20] if len(row) > 20 else None,
                "note": row[21] if len(row) > 21 else None,
                "points": row[22] if len(row) > 22 else None,
                "credit": row[23] if len(row) > 23 else None,
                "tags": row[24] if len(row) > 24 else None,
                "customer_type": row[25] if len(row) > 25 else None,
                "source": row[26] if len(row) > 26 else None,
                "priority": row[27] if len(row) > 27 else None,
                "number_of_price_of_sale": row[28] if len(row) > 28 else None,
            })

        return data
    except Exception as e:
        return e


def create_customer_from_dict(customer_data, company=None, branch=None)->bool:
    """
    Create a customer in the database use the dictionary of data that read from the file excel
    """
    try:
        #Normalizes values
        is_company = Plus.to_bool(customer_data.get("this_customer_is_a_company", ""))

        #search if exist this customer
        customer_type = None
        type_name = customer_data.get("customer_type")
        if type_name:

            #get the type customer but if not exist we will to create the type customer
            customer_type, _ = CustomerType.objects.get_or_create(
                company=company,
                name=type_name.strip(),
                defaults={"color": "#3498db"}
            )

        #search user source if exist and if not exist in the database of the company we will to create
        source = None
        source_name = customer_data.get("source")
        if source_name:
            source, _ = CustomerSource.objects.get_or_create(
                company=company,
                name=source_name.strip()
            )

        #Create the customer
        customer = Customer.objects.create(
            company=company,
            branch=branch,
            name=customer_data.get("name"),
            email=customer_data.get("email"),
            phone=customer_data.get("phone"),
            cellphone=customer_data.get("cellphone"),
            date_of_birth=customer_data.get("date_of_birth") or None,
            gender=customer_data.get("gender"),
            country=customer_data.get("country"),
            address=customer_data.get("address"),
            city=customer_data.get("city"),
            state=customer_data.get("state"),
            postal_code=customer_data.get("postal_code"),
            num_ext=customer_data.get("num_ext"),
            num_int=customer_data.get("num_int"),
            reference=customer_data.get("reference"),
            this_customer_is_a_company=is_company,
            company_name=customer_data.get("company_name"),
            contact_name=customer_data.get("contact_name"),
            contact_email=customer_data.get("contact_email"),
            contact_phone=customer_data.get("contact_phone"),
            contact_cellphone=customer_data.get("contact_cellphone"),
            website=customer_data.get("website"),
            note=customer_data.get("note"),
            points=customer_data.get("points") or 0,
            credit=customer_data.get("credit") or 0,
            tags=customer_data.get("tags") or {},
            customer_type=customer_type,
            source=source,
            priority=customer_data.get("priority") or 0,
            number_of_price_of_sale=customer_data.get("number_of_price_of_sale") or 1,
        )

        return True

    except ValidationError as e:
        return False
    except Exception as e:
        return False
    
def upload_customers_with_excel(user, excel_file)->bool:
    data_customers=get_the_data_of_the_file_excel(excel_file) #read the file excel and get his data like a list

    #when already have all the information, now we will save all the customers in the database
    customers_that_not_was_add=''
    message_answer='customers.message.al-the-customer-was-added'
    for customer in data_customers:
        answer=create_customer_from_dict(customer, user.company, user.branch)

        #we will see if can added this customer, if not can add this customer save his information and his error for show after in the frontend
        all_the_customer_was_added=True
        if not answer:
            all_the_customer_was_added=False
            message_answer='customers.message.not-all-the-customer-was-added'
            customers_that_not_was_add+=customer["name"]+','


    return {
        "success": all_the_customer_was_added,
        "answer": message_answer,
        "error": customers_that_not_was_add
    }  