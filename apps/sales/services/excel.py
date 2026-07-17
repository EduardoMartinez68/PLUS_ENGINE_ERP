import openpyxl
from decimal import Decimal
from django.db import transaction
from django.db.models import Q
from datetime import datetime
from django.utils import timezone

from apps.sales.models import (
    Sale,
    SaleItem,
    SaleHistory,
    SalePaymentMethod
)

from apps.customers.models import Customer
from core.models import CustomUser


# =========================================================
# HELPERS
# =========================================================
def parse_datetime(value):
    """
    Convert datetime safely with timezone support
    """

    if not value:
        return None

    # already datetime
    if isinstance(value, datetime):

        if timezone.is_naive(value):
            return timezone.make_aware(
                value,
                timezone.get_current_timezone()
            )

        return value

    # string datetime
    try:

        dt = datetime.strptime(
            str(value),
            "%Y-%m-%d %H:%M:%S"
        )

        return timezone.make_aware(
            dt,
            timezone.get_current_timezone()
        )

    except:
        return None
    
def this_sales_is_valid(valor):
    # 2. Lo convertimos a string, quitamos espacios y lo pasamos a minúsculas
    valor_limpio = str(valor).strip().lower()

    # 3. Definimos los casos de texto inválidos
    valores_invalidos = ['', 'none', 'null', '-']

    # 4. Comprobamos si es un valor válido
    return valor_limpio not in valores_invalidos


def safe_value(value):
    """
    Clean values
    """
    if value is None:
        return ""

    return str(value).strip()


def safe_decimal(value):
    """
    Convert decimal safely
    """
    try:
        if value is None or value == "":
            return Decimal("0.00")

        value = str(value).replace(",", "").strip()

        return Decimal(value)

    except:
        return Decimal("0.00")


def normalize_text(value):
    """
    Normalize text for comparisons
    """
    return safe_value(value).lower()

def get_value(row, headers, key):
    index = headers.get(normalize_text(key))

    if index is None:
        print(f"HEADER NOT FOUND -> {key}")
        return None

    return row[index]
# =========================================================
# READ EXCEL
# =========================================================

def get_excel_data(excel_file):
    """
    Read all sheets from excel
    """

    wb = openpyxl.load_workbook(excel_file)

    sales_sheet = wb["Sales"]
    items_sheet = wb["sale_items"]
    payments_sheet = wb["payments"]

    # -----------------------------------------------------
    # SALES
    # -----------------------------------------------------

    sales_headers = {}

    for idx, cell in enumerate(sales_sheet[1]):
        sales_headers[normalize_text(cell.value)] = idx

    sales_data = []
    for row in sales_sheet.iter_rows(min_row=2, values_only=True):

        sales_data.append({
            "reference": row[0],
            "title": row[1],
            "customer_name": row[2],
            "doctor_name": row[3],
            "subtotal": row[4],
            "discount_total": row[5],
            "tax_total": row[6],
            "total": row[7],
            "amount_paid": row[8],
            "balance": row[9],
            "currency": row[10],
            "status": row[11],
            "creation_date": row[12],
            "expiration_date": row[13],
        })

    # -----------------------------------------------------
    # ITEMS
    # -----------------------------------------------------
    items_data = []

    for row in items_sheet.iter_rows(min_row=2, values_only=True):

        items_data.append({
            "sale_reference": row[0],
            "name": row[1],
            "quantity": row[2],
            "unit_price": row[3],
            "discount": row[4],
            "tax_rate": row[5],
            "tax_amount": row[6],
            "subtotal": row[7],
            "total": row[8],
        })

    # -----------------------------------------------------
    # PAYMENTS
    # -----------------------------------------------------
    payments_data = []

    for row in payments_sheet.iter_rows(min_row=2, values_only=True):

        payments_data.append({
            "sale_reference": row[0],
            "cash_received": row[1],
            "change_given": row[2],
            "old_amount_paid": row[3],
            "old_balance": row[4],
            "new_amount_paid": row[5],
            "new_balance": row[6],
            "date": row[7],
            "payment_method": row[8],
            "payment_amount": row[9],
        })

    return {
        "sales": sales_data,
        "items": items_data,
        "payments": payments_data
    }


# =========================================================
# CUSTOMER
# =========================================================
def get_or_create_customer(customer_name, company, branch):
    """
    Search customer by decrypted name.
    If not exists create customer.
    """

    customer_name = safe_value(customer_name)

    if not customer_name:
        return None

    customer_name_normalized = (
        customer_name.lower().strip()
    )

    customers = Customer.objects.filter(
        company=company,
        branch=branch
    )

    customer = None

    for c in customers:

        try:
            if (
                c.name and
                c.name.lower().strip() == customer_name_normalized
            ):
                customer = c
                break

        except Exception:
            pass

    # -----------------------------------------
    # CUSTOMER FOUND
    # -----------------------------------------

    if customer:
        print("CUSTOMER FOUND")
        return customer

    # -----------------------------------------
    # CREATE CUSTOMER
    # -----------------------------------------

    print("CUSTOMER CREATED")

    customer = Customer.objects.create(
        company=company,
        branch=branch,
        name=customer_name
    )

    return customer

# =========================================================
# DOCTOR
# =========================================================

def get_doctor(doctor_name, company, branch):
    """
    Search doctor.
    If not exists return None.
    """

    doctor_name = safe_value(doctor_name)

    if not doctor_name:
        return None

    doctor = CustomUser.objects.filter(
        company=company,
        branch=branch
    ).filter(
        Q(name__iexact=doctor_name) |
        Q(username__iexact=doctor_name) |
        Q(email__iexact=doctor_name)
    ).first()

    return doctor


# =========================================================
# CREATE SALE
# =========================================================
from apps.sales.services.sales import Sales

@transaction.atomic
def create_sale_from_dict(
    sale_data,
    items_data,
    payments_data,
    user
):
    try:
        """
        Create complete sale
        """

        company = user.company
        branch = user.branch

        # -----------------------------------------------------
        # CUSTOMER
        # -----------------------------------------------------

        customer = get_or_create_customer(
            sale_data.get("customer_name"),
            company,
            branch
        )
        # -----------------------------------------------------
        # DOCTOR
        # -----------------------------------------------------

        doctor_name = safe_value(
            sale_data.get("doctor_name")
        )

        doctor = get_doctor(
            doctor_name,
            company,
            branch
        )

        # -----------------------------------------------------
        # COMMENTS
        # -----------------------------------------------------

        comments = ""

        if doctor is None and doctor_name:
            comments = f"Doctor not found during migration: {doctor_name}"

        # -----------------------------------------------------
        # CREATE SALE
        # -----------------------------------------------------
        sale = Sale.objects.create(
            reference=sale_data.get("reference"),
            company=company,
            branch=branch,

            customer=customer,
            user=doctor,

            name_sale=f"{sale_data.get('title','Venta')} {'' if doctor else f'- {doctor_name}'}",

            subtotal=safe_decimal(sale_data.get("subtotal")),
            discount_total=safe_decimal(sale_data.get("discount_total")),
            tax_total=safe_decimal(sale_data.get("tax_total")),
            total=safe_decimal(sale_data.get("total")),
            amount_paid=safe_decimal(sale_data.get("amount_paid")),
            balance=safe_decimal(sale_data.get("balance")),

            currency=safe_value(
                sale_data.get("currency")
            ) or "MXN",

            status=safe_value(
                sale_data.get("status")
            ) or "pending",

            note_cancelled=comments,

            created_by=user,
            expiration_date=parse_datetime(sale_data.get("expiration_date")),
            payment_term_days=sale_data.get("payment_term_days", 1) #this is for that can be buy after
        )

        #when create the sale now we will update of date of the sale for that have chronology
        creation_date_sale=parse_datetime(sale_data.get("creation_date"))
        if creation_date_sale:
            sale.creationDate = creation_date_sale
            sale.save(update_fields=["creationDate"])

        # -----------------------------------------------------
        # ITEMS
        # -----------------------------------------------------

        for item_data in items_data:

            if safe_value(
                item_data.get("sale_reference")
            ) != safe_value(
                sale_data.get("reference")
            ):
                continue

            SaleItem.objects.create(
                sale=sale,

                name=safe_value(
                    item_data.get("name")
                ),

                quantity=safe_decimal(
                    item_data.get("quantity")
                ),

                unit_price=safe_decimal(
                    item_data.get("unit_price")
                ),

                discount=safe_decimal(
                    item_data.get("discount")
                ),

                tax_rate=safe_decimal(
                    item_data.get("tax_rate")
                ),

                tax_amount=safe_decimal(
                    item_data.get("tax_amount")
                ),

                subtotal=safe_decimal(
                    item_data.get("subtotal")
                ),

                total=safe_decimal(
                    item_data.get("total")
                ),

                taxes=[]
            )

        # -----------------------------------------------------
        # PAYMENTS
        # -----------------------------------------------------
        for payment_data in payments_data:
            if safe_value(
                payment_data.get("sale_reference")
            ) != safe_value(
                sale_data.get("reference")
            ):
                continue
            
            reference_move=payment_data.get("sale_reference")
            payment = SaleHistory.objects.create(
                sale=sale,

                cash_received=safe_decimal(
                    payment_data.get("cash_received")
                ),

                change_given=safe_decimal(
                    payment_data.get("change_given")
                ),

                old_amount_paid=safe_decimal(
                    payment_data.get("old_amount_paid")
                ),

                old_balance=safe_decimal(
                    payment_data.get("old_balance")
                ),

                new_amount_paid=safe_decimal(
                    payment_data.get("new_amount_paid")
                ),

                new_balance=safe_decimal(
                    payment_data.get("new_balance")
                ),

                created_by=user
            )
            

            if creation_date_sale:
                payment.date = creation_date_sale
                payment.save(update_fields=["date"])

            # ---------------------------------------------
            # PAYMENT METHOD
            # ---------------------------------------------

            payment_method =SalePaymentMethod.objects.create(
                company=company,
                branch=branch,

                sale=sale,
                payment=payment,

                user=user,

                method=safe_value(
                    payment_data.get("payment_method")
                ) or "cash",

                amount=safe_decimal(
                    payment_data.get("payment_amount")
                ),

                comment=f'Venta: {reference_move}-Migración de datos'
            )

            Sales.add_money_movement(
                company=company,
                branch=branch,
                amount=safe_decimal(
                    payment_data.get("payment_amount")
                )
            )
    
            if creation_date_sale:
                payment_method.date = creation_date_sale
                payment_method.save(update_fields=["date"])
        return sale
    except Exception as e:
        print(e)
        return None


# =========================================================
# MAIN IMPORT
# =========================================================
        
def upload_excel(user, excel_file):
    """
    Upload sales from excel
    """
    try:

        excel_data = get_excel_data(excel_file)

        sales = excel_data["sales"]
        items = excel_data["items"]
        payments = excel_data["payments"]

        imported_sales = 0
        errors = []
        for sale_data in sales:
            if not this_sales_is_valid(sale_data.get('reference', '')):
                continue

            try:

                reference = safe_value(
                    sale_data.get("reference")
                )

                # -----------------------------------------
                # AVOID DUPLICATES
                # -----------------------------------------

                sale_exists = Sale.objects.filter(
                    company=user.company,
                    branch=user.branch,
                    reference=reference
                ).exists()

                if sale_exists:
                    errors.append(
                        f"Sale already exists: {reference}"
                    )
                    continue

                # -----------------------------------------
                # CREATE SALE
                # -----------------------------------------
                create_sale_from_dict(
                    sale_data=sale_data,
                    items_data=items,
                    payments_data=payments,
                    user=user
                )

                imported_sales += 1
            except Exception as e:
                print(e)
                errors.append(
                    f"{reference}: {str(e)}"
                )

        return {
            "success": len(errors) == 0,
            "imported_sales": imported_sales,
            "errors": errors
        }

    except Exception as e:
        print(e)
        return {
            "success": False,
            "imported_sales": 0,
            "errors": [str(e)]
        }